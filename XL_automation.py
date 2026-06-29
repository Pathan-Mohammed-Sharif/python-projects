from openpyxl import Workbook
from openpyxl.styles import Font

# Create workbook
wb = Workbook()
sheet = wb.active
sheet.title = "Monthly Report"

# Heading
sheet["A1"] = "Employee"
sheet["B1"] = "Sales"

sheet["A1"].font = Font(bold=True)
sheet["B1"].font = Font(bold=True)

# Sample data
data = [
    ["Mohammed", 25000],
    ["Rahul", 18000],
    ["Anjali", 32000],
    ["Ravi", 21000],
    ["Sneha", 27000]
]

# Add data
row = 2
for employee, sales in data:
    sheet.cell(row=row, column=1).value = employee
    sheet.cell(row=row, column=2).value = sales
    row += 1

# Calculations
sales_list = [x[1] for x in data]

total = sum(sales_list)
average = total / len(sales_list)
highest = max(sales_list)
lowest = min(sales_list)

# Report
sheet["D2"] = "Total Sales"
sheet["E2"] = total

sheet["D3"] = "Average Sales"
sheet["E3"] = average

sheet["D4"] = "Highest Sales"
sheet["E4"] = highest

sheet["D5"] = "Lowest Sales"
sheet["E5"] = lowest

# Save file
wb.save("Monthly_Report.xlsx")

print("Monthly Report Generated Successfully!")
