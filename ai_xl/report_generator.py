from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import PieChart, Reference


class ReportGenerator:

    def _init_(self, workbook, sheet):
        self.workbook = workbook
        self.sheet = sheet

    def format_sheet(self):

        header_fill = PatternFill(
            start_color="4F81BD",
            end_color="4F81BD",
            fill_type="solid"
        )

        summary_fill = PatternFill(
            start_color="D9EAD3",
            end_color="D9EAD3",
            fill_type="solid"
        )

        for cell in self.sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row in range(2, 8):
            self.sheet[f"D{row}"].font = Font(bold=True)
            self.sheet[f"D{row}"].fill = summary_fill

        self.sheet.column_dimensions["A"].width = 25
        self.sheet.column_dimensions["B"].width = 18
        self.sheet.column_dimensions["D"].width = 25
        self.sheet.column_dimensions["E"].width = 20

    def write_ai_summary(self, summary):

        start_row = 10

        self.sheet[f"A{start_row}"] = "AI SALES ANALYSIS"
        self.sheet[f"A{start_row}"].font = Font(
            bold=True,
            size=14
        )

        row = start_row + 1

        for line in summary.split("\n"):

            self.sheet.cell(
                row=row,
                column=1
            ).value = line

            row += 1

    def create_pie_chart(self, employee_count):

        labels = Reference(
            self.sheet,
            min_col=1,
            min_row=2,
            max_row=employee_count + 1
        )

        data = Reference(
            self.sheet,
            min_col=2,
            min_row=1,
            max_row=employee_count + 1
        )

        pie = PieChart()

        pie.add_data(
            data,
            titles_from_data=True
        )

        pie.set_categories(labels)

        pie.title = "Sales Distribution"

        self.sheet.add_chart(
            pie,
            "G20"
        )

    def highlight_top_bottom(self, data):

        highest = max([x[1] for x in data])
        lowest = min([x[1] for x in data])

        green = PatternFill(
            start_color="90EE90",
            end_color="90EE90",
            fill_type="solid"
        )

        red = PatternFill(
            start_color="FF9999",
            end_color="FF9999",
            fill_type="solid"
        )

        row = 2

        for employee, sales in data:

            if sales == highest:
                self.sheet[f"B{row}"].fill = green

            elif sales == lowest:
                self.sheet[f"B{row}"].fill = red

            row += 1

    def save_report(
        self,
        filename="AI_Monthly_Report.xlsx"
    ):

        self.workbook.save(filename)

        print("\n================================")
        print(" AI REPORT GENERATED SUCCESSFULLY")
        print("================================")
        print(f"Saved as: {filename}")

    def generate_complete_report(
        self,
        data,
        summary
    ):

        self.format_sheet()

        self.write_ai_summary(summary)

        self.highlight_top_bottom(data)

        self.create_pie_chart(len(data))

        self.save_report()
