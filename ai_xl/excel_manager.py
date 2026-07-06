from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference


class ExcelManager:

    def _init_(self):
        self.workbook = None
        self.sheet = None
        self.data = []

    def create_new_sheet(self):

        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "Monthly Report"

        self.sheet["A1"] = "Employee"
        self.sheet["B1"] = "Sales"

        self.sheet["A1"].font = Font(bold=True)
        self.sheet["B1"].font = Font(bold=True)

        employees = int(input("Enter number of employees: "))

        self.data = []

        for i in range(employees):

            print(f"\nEmployee {i+1}")

            name = input("Name : ")
            sales = float(input("Sales : "))

            self.data.append([name, sales])

            self.sheet.cell(row=i + 2, column=1).value = name
            self.sheet.cell(row=i + 2, column=2).value = sales

        return self.data

    def import_sheet(self):

        path = input("Enter Excel File Path : ")

        self.workbook = load_workbook(path)
        self.sheet = self.workbook.active

        self.data = []

        row = 2

        while True:

            name = self.sheet.cell(row=row, column=1).value
            sales = self.sheet.cell(row=row, column=2).value

            if name is None:
                break

            self.data.append([name, float(sales)])

            row += 1

        return self.data

    def calculate_statistics(self):

        sales = [x[1] for x in self.data]

        total = sum(sales)
        average = total / len(sales)
        highest = max(sales)
        lowest = min(sales)

        highest_employee = ""

        lowest_employee = ""

        for emp, value in self.data:

            if value == highest:
                highest_employee = emp

            if value == lowest:
                lowest_employee = emp

        self.sheet["D2"] = "Total Sales"
        self.sheet["E2"] = total

        self.sheet["D3"] = "Average Sales"
        self.sheet["E3"] = average

        self.sheet["D4"] = "Highest Sales"
        self.sheet["E4"] = highest

        self.sheet["D5"] = "Lowest Sales"
        self.sheet["E5"] = lowest

        self.sheet["D6"] = "Top Performer"
        self.sheet["E6"] = highest_employee

        self.sheet["D7"] = "Lowest Performer"
        self.sheet["E7"] = lowest_employee

        return {
            "total": total,
            "average": average,
            "highest": highest,
            "lowest": lowest,
            "highest_employee": highest_employee,
            "lowest_employee": lowest_employee
        }

    def create_chart(self):

        chart = BarChart()

        data = Reference(
            self.sheet,
            min_col=2,
            min_row=1,
            max_row=len(self.data) + 1
        )

        categories = Reference(
            self.sheet,
            min_col=1,
            min_row=2,
            max_row=len(self.data) + 1
        )

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        chart.title = "Employee Sales"
        chart.y_axis.title = "Sales"
        chart.x_axis.title = "Employees"

        self.sheet.add_chart(chart, "G2")

    def save(self, filename="Monthly_Report.xlsx"):

        self.workbook.save(filename)

        print(f"\nReport saved as {filename}")

    def get_data(self):
        return self.data

    def get_sheet(self):
        return self.sheet

    def get_workbook(self):
        return self.workbook
